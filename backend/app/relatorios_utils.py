from datetime import datetime
from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

# Paleta -- mesma cor primária usada no frontend (Tailwind indigo-600), para
# os relatórios terem a mesma identidade visual do sistema.
COR_PRIMARIA = colors.HexColor("#4f46e5")
COR_PRIMARIA_ESCURA = colors.HexColor("#3730a3")
COR_ZEBRA = colors.HexColor("#f5f5f8")
COR_BORDA = colors.HexColor("#e0e0e6")
COR_TEXTO_CLARO = colors.HexColor("#6b7280")

COR_PRIMARIA_HEX = "4F46E5"
COR_ZEBRA_HEX = "F5F5F8"
COR_BORDA_HEX = "D9D9E3"

# Cabeçalhos que indicam coluna monetária -- usados para alinhar à direita e
# aplicar formatação de moeda automaticamente, sem precisar marcar cada
# relatório manualmente.
_PALAVRAS_MONETARIAS = ("r$", "valor", "custo", "total")


def _eh_coluna_monetaria(nome_coluna: str) -> bool:
    nome = nome_coluna.lower()
    return any(p in nome for p in _PALAVRAS_MONETARIAS)


def _rodape_com_pagina(canvas, doc):
    canvas.saveState()
    largura, _ = landscape(A4)
    canvas.setStrokeColor(COR_BORDA)
    canvas.setLineWidth(0.5)
    canvas.line(1.5 * cm, 1.3 * cm, largura - 1.5 * cm, 1.3 * cm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(COR_TEXTO_CLARO)
    canvas.drawString(1.5 * cm, 0.9 * cm, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Controle de Frota")
    canvas.drawRightString(largura - 1.5 * cm, 0.9 * cm, f"Página {doc.page}")
    canvas.restoreState()


def gerar_pdf(titulo: str, colunas: list[str], linhas: list[list], subtitulo: str = "") -> BytesIO:
    """Gera um PDF de tabela, em paisagem, com cabeçalho tipo timbre, colunas monetárias
    alinhadas à direita e numeração de página no rodapé."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.3 * cm, bottomMargin=1.8 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloRelatorio", parent=estilos["Title"],
        textColor=COR_PRIMARIA_ESCURA, fontSize=18, spaceAfter=2, alignment=TA_LEFT,
    )
    estilo_subtitulo = ParagraphStyle(
        "SubtituloRelatorio", parent=estilos["Normal"],
        textColor=COR_TEXTO_CLARO, fontSize=10,
    )
    estilo_marca = ParagraphStyle(
        "Marca", parent=estilos["Normal"],
        textColor=COR_TEXTO_CLARO, fontSize=8, alignment=TA_RIGHT,
    )

    elementos = [
        Table(
            [[Paragraph(titulo, estilo_titulo), Paragraph("CONTROLE DE FROTA", estilo_marca)]],
            colWidths=[doc.width * 0.7, doc.width * 0.3],
        ),
    ]
    elementos[0].setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    if subtitulo:
        elementos.append(Paragraph(subtitulo, estilo_subtitulo))

    # linha de destaque abaixo do cabeçalho, reforçando a identidade visual
    linha_destaque = Table([[""]], colWidths=[doc.width], rowHeights=[0.08 * cm])
    linha_destaque.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), COR_PRIMARIA)]))
    elementos.append(Spacer(1, 0.3 * cm))
    elementos.append(linha_destaque)
    elementos.append(Spacer(1, 0.5 * cm))

    colunas_monetarias = [i for i, c in enumerate(colunas) if _eh_coluna_monetaria(c)]

    def _formatar_celula(valor, indice_coluna):
        if valor is None:
            return "—"
        if indice_coluna in colunas_monetarias and isinstance(valor, (int, float)):
            return f"R$ {valor:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
        return str(valor)

    dados_tabela = [colunas] + [
        [_formatar_celula(c, i) for i, c in enumerate(linha)] for linha in linhas
    ]

    # detecta se a última linha é um total (primeira célula contém "Total") para destacar
    ultima_e_total = bool(linhas) and isinstance(linhas[-1][0], str) and "total" in linhas[-1][0].lower()

    estilo_tabela = [
        ("BACKGROUND", (0, 0), (-1, 0), COR_PRIMARIA),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COR_ZEBRA]),
        ("LINEBELOW", (0, 0), (-1, 0), 1, COR_PRIMARIA_ESCURA),
        ("LINEBELOW", (0, 1), (-1, -2 if ultima_e_total else -1), 0.4, COR_BORDA),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]
    for indice_coluna in colunas_monetarias:
        estilo_tabela.append(("ALIGN", (indice_coluna, 0), (indice_coluna, -1), "RIGHT"))
    if ultima_e_total:
        estilo_tabela.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
        estilo_tabela.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2ff")))
        estilo_tabela.append(("LINEABOVE", (0, -1), (-1, -1), 1, COR_PRIMARIA))

    # larguras proporcionais ao maior conteúdo de cada coluna, preenchendo a página toda
    # (em vez de deixar o reportlab encolher a tabela pro tamanho mínimo do conteúdo)
    pesos = []
    for i, col in enumerate(colunas):
        maior = len(str(col))
        for linha in dados_tabela[1:]:
            maior = max(maior, len(str(linha[i])))
        pesos.append(maior + 2)
    soma_pesos = sum(pesos) or 1
    larguras = [doc.width * (p / soma_pesos) for p in pesos]

    tabela = Table(dados_tabela, repeatRows=1, colWidths=larguras)
    tabela.setStyle(TableStyle(estilo_tabela))
    elementos.append(tabela)

    doc.build(elementos, onFirstPage=_rodape_com_pagina, onLaterPages=_rodape_com_pagina)
    buffer.seek(0)
    return buffer


def gerar_excel(titulo: str, colunas: list[str], linhas: list[list]) -> BytesIO:
    """Gera um .xlsx com título, cabeçalho estilizado, formatação de moeda nas colunas
    monetárias, tabela nomeada (com autofiltro) e cabeçalho congelado."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # limite do Excel para nome de aba

    num_colunas = len(colunas)
    ultima_coluna_letra = get_column_letter(num_colunas)

    # título mesclado na primeira linha
    ws.merge_cells(f"A1:{ultima_coluna_letra}1")
    celula_titulo = ws["A1"]
    celula_titulo.value = titulo
    celula_titulo.font = Font(bold=True, size=14, color=COR_PRIMARIA_HEX)
    celula_titulo.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{ultima_coluna_letra}2")
    celula_data = ws["A2"]
    celula_data.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Controle de Frota"
    celula_data.font = Font(size=9, italic=True, color="6B7280")
    ws.row_dimensions[2].height = 16

    linha_cabecalho = 4
    for indice, nome_coluna in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=indice, value=nome_coluna)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color=COR_PRIMARIA_HEX, end_color=COR_PRIMARIA_HEX, fill_type="solid")
        celula.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha_cabecalho].height = 20

    colunas_monetarias = {i for i, c in enumerate(colunas, start=1) if _eh_coluna_monetaria(c)}
    borda_fina = Border(*(Side(style="thin", color=COR_BORDA_HEX) for _ in range(4)))

    linha_atual = linha_cabecalho + 1
    for linha in linhas:
        for indice, valor in enumerate(linha, start=1):
            celula = ws.cell(row=linha_atual, column=indice, value=("—" if valor is None else valor))
            celula.border = borda_fina
            if indice in colunas_monetarias and isinstance(valor, (int, float)):
                celula.number_format = 'R$ #,##0.00'
                celula.alignment = Alignment(horizontal="right")
        linha_atual += 1

    linha_final = linha_atual - 1
    if linha_final >= linha_cabecalho:
        tabela_excel = ExcelTable(
            displayName="TabelaRelatorio",
            ref=f"A{linha_cabecalho}:{ultima_coluna_letra}{linha_final}",
        )
        tabela_excel.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tabela_excel)

    for indice, nome_coluna in enumerate(colunas, start=1):
        letra = get_column_letter(indice)
        valores_coluna = [nome_coluna] + [str(linha[indice - 1]) for linha in linhas if linha[indice - 1] is not None]
        largura_maxima = max((len(v) for v in valores_coluna), default=10)
        ws.column_dimensions[letra].width = min(largura_maxima + 4, 50)

    ws.freeze_panes = f"A{linha_cabecalho + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def resposta_pdf(buffer: BytesIO, nome_arquivo: str) -> StreamingResponse:
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


def resposta_excel(buffer: BytesIO, nome_arquivo: str) -> StreamingResponse:
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )
