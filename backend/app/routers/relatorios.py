from datetime import date
from typing import Optional

from fastapi import APIRouter, HTTPException, Query

from app.database import get_supabase
from app.relatorios_utils import gerar_pdf, gerar_excel, resposta_pdf, resposta_excel

router = APIRouter(prefix="/relatorios", tags=["Relatórios"])


# ==================== Histórico Completo de Máquina ====================

def _montar_historico(maquina_id: int):
    sb = get_supabase()

    maquina = sb.table("maquinas").select("*").eq("id", maquina_id).execute()
    if not maquina.data:
        raise HTTPException(status_code=404, detail="Máquina não encontrada")
    maquina = maquina.data[0]

    manutencoes = (
        sb.table("manutencoes").select("data, tipo, descricao, custo, horimetro, km")
        .eq("maquina_id", maquina_id).order("data", desc=True).execute().data
    )
    abastecimentos = (
        sb.table("abastecimentos").select("data, tipo_combustivel, quantidade, valor_total")
        .eq("maquina_id", maquina_id).order("data", desc=True).execute().data
    )
    preventivas = (
        sb.table("planos_preventiva").select("descricao, proxima_data, proximo_horimetro, proximo_km")
        .eq("maquina_id", maquina_id).execute().data
    )
    checklists = (
        sb.table("checklist_execucoes").select("data, operador, tem_pendencia")
        .eq("maquina_id", maquina_id).order("data", desc=True).limit(20).execute().data
    )

    return maquina, manutencoes, abastecimentos, preventivas, checklists


@router.get("/historico-maquina/{maquina_id}/pdf")
def historico_maquina_pdf(maquina_id: int):
    maquina, manutencoes, abastecimentos, preventivas, checklists = _montar_historico(maquina_id)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Histórico Completo — {maquina['codigo']}", estilos["Title"]),
        Paragraph(f"{maquina['marca'] or ''} {maquina['modelo'] or ''} — {maquina['tipo']}", estilos["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    def _secao(nome, colunas, linhas):
        elementos.append(Paragraph(nome, estilos["Heading2"]))
        if not linhas:
            elementos.append(Paragraph("Nenhum registro.", estilos["Normal"]))
        else:
            dados = [colunas] + [[str(c) if c is not None else "—" for c in linha] for linha in linhas]
            tabela = Table(dados, repeatRows=1)
            tabela.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f7")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
            ]))
            elementos.append(tabela)
        elementos.append(Spacer(1, 0.4 * cm))

    _secao("Manutenções", ["Data", "Tipo", "Descrição", "Custo", "Horímetro", "Km"],
           [[m["data"], m["tipo"], m["descricao"], m["custo"], m["horimetro"], m["km"]] for m in manutencoes])
    _secao("Abastecimentos", ["Data", "Combustível", "Quantidade", "Valor"],
           [[a["data"], a["tipo_combustivel"], a["quantidade"], a["valor_total"]] for a in abastecimentos])
    _secao("Planos de Preventiva", ["Descrição", "Próxima Data", "Próximo Horímetro", "Próximo Km"],
           [[p["descricao"], p["proxima_data"], p["proximo_horimetro"], p["proximo_km"]] for p in preventivas])
    _secao("Últimos Checklists", ["Data", "Operador", "Pendência?"],
           [[c["data"][:16].replace("T", " "), c["operador"], "Sim" if c["tem_pendencia"] else "Não"] for c in checklists])

    doc.build(elementos)
    buffer.seek(0)
    return resposta_pdf(buffer, f"historico-{maquina['codigo']}.pdf")


@router.get("/historico-maquina/{maquina_id}/excel")
def historico_maquina_excel(maquina_id: int):
    maquina, manutencoes, abastecimentos, preventivas, checklists = _montar_historico(maquina_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = Workbook()

    def _aba(nome, colunas, linhas, primeira=False):
        ws = wb.active if primeira else wb.create_sheet(nome)
        if not primeira:
            ws.title = nome
        ws.append(colunas)
        for celula in ws[1]:
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for linha in linhas:
            ws.append(["—" if c is None else c for c in linha])

    _aba("Manutenções", ["Data", "Tipo", "Descrição", "Custo", "Horímetro", "Km"],
         [[m["data"], m["tipo"], m["descricao"], m["custo"], m["horimetro"], m["km"]] for m in manutencoes], primeira=True)
    _aba("Abastecimentos", ["Data", "Combustível", "Quantidade", "Valor"],
         [[a["data"], a["tipo_combustivel"], a["quantidade"], a["valor_total"]] for a in abastecimentos])
    _aba("Preventivas", ["Descrição", "Próxima Data", "Próximo Horímetro", "Próximo Km"],
         [[p["descricao"], p["proxima_data"], p["proximo_horimetro"], p["proximo_km"]] for p in preventivas])
    _aba("Checklists", ["Data", "Operador", "Pendência?"],
         [[c["data"][:16].replace("T", " "), c["operador"], "Sim" if c["tem_pendencia"] else "Não"] for c in checklists])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return resposta_excel(buffer, f"historico-{maquina['codigo']}.xlsx")


# ==================== Custos Gerais ====================

def _montar_custos_gerais(data_inicio: Optional[date], data_fim: Optional[date], centro_despesa_id: Optional[int]):
    sb = get_supabase()

    maquinas_query = sb.table("maquinas").select("id, codigo, centro_despesa_id, centros_despesa(nome)")
    if centro_despesa_id:
        maquinas_query = maquinas_query.eq("centro_despesa_id", centro_despesa_id)
    maquinas = maquinas_query.execute().data
    for m in maquinas:
        m["centro_despesa_nome"] = (m.pop("centros_despesa", None) or {}).get("nome")
    maquinas_por_id = {m["id"]: m for m in maquinas}
    ids_permitidos = set(maquinas_por_id.keys())

    manutencoes_query = sb.table("manutencoes").select("maquina_id, data, custo")
    if data_inicio:
        manutencoes_query = manutencoes_query.gte("data", data_inicio.isoformat())
    if data_fim:
        manutencoes_query = manutencoes_query.lte("data", data_fim.isoformat())
    manutencoes = [m for m in manutencoes_query.execute().data if m["maquina_id"] in ids_permitidos]

    abastecimentos_query = sb.table("abastecimentos").select("maquina_id, data, valor_total")
    if data_inicio:
        abastecimentos_query = abastecimentos_query.gte("data", data_inicio.isoformat())
    if data_fim:
        abastecimentos_query = abastecimentos_query.lte("data", data_fim.isoformat())
    abastecimentos = [a for a in abastecimentos_query.execute().data if a["maquina_id"] in ids_permitidos]

    linhas = []
    for m in manutencoes:
        maquina = maquinas_por_id[m["maquina_id"]]
        linhas.append([maquina["codigo"], maquina["centro_despesa_nome"], "Manutenção", m["data"], m["custo"] or 0])
    for a in abastecimentos:
        maquina = maquinas_por_id[a["maquina_id"]]
        linhas.append([maquina["codigo"], maquina["centro_despesa_nome"], "Combustível", a["data"], a["valor_total"] or 0])

    linhas.sort(key=lambda linha: linha[3], reverse=True)
    return linhas


@router.get("/custos-gerais/pdf")
def custos_gerais_pdf(
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    centro_despesa_id: Optional[int] = None,
):
    linhas = _montar_custos_gerais(data_inicio, data_fim, centro_despesa_id)
    total = sum(linha[4] for linha in linhas)
    subtitulo = f"Total: R$ {total:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    buffer = gerar_pdf(
        "Custos Gerais",
        ["Máquina", "Centro de Despesa", "Categoria", "Data", "Valor (R$)"],
        linhas,
        subtitulo=subtitulo,
    )
    return resposta_pdf(buffer, "custos-gerais.pdf")


@router.get("/custos-gerais/excel")
def custos_gerais_excel(
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    centro_despesa_id: Optional[int] = None,
):
    linhas = _montar_custos_gerais(data_inicio, data_fim, centro_despesa_id)
    buffer = gerar_excel(
        "Custos Gerais",
        ["Máquina", "Centro de Despesa", "Categoria", "Data", "Valor (R$)"],
        linhas,
    )
    return resposta_excel(buffer, "custos-gerais.xlsx")


# ==================== Preventivas Vencidas e a Vencer ====================

def _montar_preventivas_relatorio():
    from app.routers.preventivas import _calcular_status

    sb = get_supabase()
    planos = sb.table("planos_preventiva").select("*").execute().data
    maquina_ids = list({p["maquina_id"] for p in planos})
    maquinas = (
        sb.table("maquinas").select("id, codigo, horimetro_atual, km_atual").in_("id", maquina_ids).execute().data
        if maquina_ids else []
    )
    maquinas_por_id = {m["id"]: m for m in maquinas}

    linhas = []
    for plano in planos:
        maquina = maquinas_por_id.get(plano["maquina_id"], {})
        status = _calcular_status(plano, maquina)
        if status == "em_dia":
            continue
        linhas.append([
            maquina.get("codigo"),
            plano["descricao"],
            "Vencida" if status == "vencida" else "Próxima",
            plano.get("proxima_data"),
            plano.get("proximo_horimetro"),
            plano.get("proximo_km"),
        ])

    linhas.sort(key=lambda linha: linha[2])  # vencidas primeiro (V < P alfabeticamente... ajustamos:)
    linhas.sort(key=lambda linha: 0 if linha[2] == "Vencida" else 1)
    return linhas


@router.get("/preventivas-vencimento/pdf")
def preventivas_vencimento_pdf():
    linhas = _montar_preventivas_relatorio()
    buffer = gerar_pdf(
        "Preventivas Vencidas e a Vencer",
        ["Máquina", "Descrição", "Status", "Próxima Data", "Próximo Horímetro", "Próximo Km"],
        linhas,
    )
    return resposta_pdf(buffer, "preventivas-vencimento.pdf")


@router.get("/preventivas-vencimento/excel")
def preventivas_vencimento_excel():
    linhas = _montar_preventivas_relatorio()
    buffer = gerar_excel(
        "Preventivas",
        ["Máquina", "Descrição", "Status", "Próxima Data", "Próximo Horímetro", "Próximo Km"],
        linhas,
    )
    return resposta_excel(buffer, "preventivas-vencimento.xlsx")
